
# Para subir archivo tipo foto al servidor
from werkzeug.utils import secure_filename
import uuid  # Modulo de python para crear un string

from conexion.conexionBD import connectionBD  # Conexión a BD

import datetime
import re
import os

from os import remove  # Modulo  para remover archivo
from os import path  # Modulo para obtener la ruta o directorio


import openpyxl  # Para generar el excel
# biblioteca o modulo send_file para forzar la descarga
from flask import send_file


# Registrar una película
def procesar_form_pelicula(dataForm, poster_pelicula):
    try:
        # Procesar y guardar el póster
        poster_file = procesar_imagen_poster(poster_pelicula)
        if not poster_file:
            print("Error: No se pudo procesar el archivo del póster.")
            return None

        # Procesar presupuesto eliminando caracteres no numéricos
        presupuesto_raw = dataForm.get('presupuesto', '0')
        presupuesto_clean = re.sub(r'[^\d]', '', presupuesto_raw)  # Elimina todo excepto números
        presupuesto = int(presupuesto_clean) if presupuesto_clean else 0

        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                sql = """
                INSERT INTO tbl_peliculas (titulo_pelicula, categoria_pelicula, ano_estreno, duracion_minutos,
                descripcion, director, poster_pelicula, presupuesto)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                """
                valores = (
                    dataForm.get('titulo_pelicula'), dataForm.get('categoria_pelicula'),
                    dataForm.get('ano_estreno'), dataForm.get('duracion_minutos'),
                    dataForm.get('descripcion'), dataForm.get('director'),
                    poster_file, presupuesto
                )

                # Imprime datos para depuración
                print("Consulta SQL:", sql)
                print("Valores:", valores)

                cursor.execute(sql, valores)
                conexion_MySQLdb.commit()
                return cursor.rowcount
    except Exception as e:
        print(f"Error al registrar la película: {e}")
        return None




# Procesar imagen de póster
def procesar_imagen_poster(poster):
    try:
        filename = secure_filename(poster.filename)
        extension = os.path.splitext(filename)[1]
        unique_name = (uuid.uuid4().hex + uuid.uuid4().hex)[:100] + extension

        basepath = os.path.abspath(os.path.dirname(__file__))
        upload_dir = os.path.join(basepath, '../static/posters_peliculas/')

        if not os.path.exists(upload_dir):
            os.makedirs(upload_dir)
            os.chmod(upload_dir, 0o755)

        upload_path = os.path.join(upload_dir, unique_name)
        poster.save(upload_path)

        return unique_name
    except Exception as e:
        print(f"Error al procesar el póster: {e}")
        return None


# Lista de peliculas
def sql_lista_peliculasBD():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = """
                SELECT id_pelicula, titulo_pelicula, categoria_pelicula, ano_estreno, duracion_minutos,
                       director, poster_pelicula, presupuesto
                FROM tbl_peliculas
                ORDER BY id_pelicula DESC
                """
                cursor.execute(querySQL)
                peliculasBD = cursor.fetchall()
        return peliculasBD
    except Exception as e:
        print(f"Error al listar películas: {e}")
        return None


# Obtener detalles de una película
def sql_detalles_peliculasBD(idPelicula):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = """
                SELECT id_pelicula, titulo_pelicula, categoria_pelicula, ano_estreno, duracion_minutos,
                       descripcion, director, poster_pelicula, presupuesto,
                       DATE_FORMAT(fecha_registro, '%Y-%m-%d %h:%i %p') AS fecha_registro
                FROM tbl_peliculas
                WHERE id_pelicula = %s
                """
                cursor.execute(querySQL, (idPelicula,))
                pelicula = cursor.fetchone()
        return pelicula
    except Exception as e:
        print(f"Error al obtener detalles de la película: {e}")
        return None


# Generar un reporte de películas
def peliculasReporte():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = """
                    SELECT 
                        p.id_pelicula,
                        p.titulo_pelicula,
                        p.categoria_pelicula,
                        p.ano_estreno,
                        p.duracion_minutos,
                        p.director,
                        p.presupuesto,
                        DATE_FORMAT(p.fecha_registro, '%d de %b %Y %h:%i %p') AS fecha_registro
                    FROM tbl_peliculas AS p
                    ORDER BY p.id_pelicula DESC
                """
                cursor.execute(querySQL)
                peliculasBD = cursor.fetchall()
        return peliculasBD
    except Exception as e:
        print(f"Error en la función peliculasReporte: {e}")
        return None



def generarReporteExcelPeliculas():
    dataPeliculas = peliculasReporte()
    wb = openpyxl.Workbook()
    hoja = wb.active

    # Agregar la fila de encabezado con los títulos
    cabeceraExcel = ("Título", "Categoría", "Año de Estreno", "Duración", "Director", "Presupuesto", "Fecha de Registro")

    hoja.append(cabeceraExcel)

    # Formato para números en moneda sin decimales
    formato_moneda = '#,##0'

    # Agregar los registros a la hoja
    for registro in dataPeliculas:
        titulo_pelicula = registro['titulo_pelicula']
        categoria_pelicula = registro['categoria_pelicula']
        ano_estreno = registro['ano_estreno']
        duracion_minutos = registro['duracion_minutos']
        director = registro['director']
        presupuesto = registro['presupuesto']
        fecha_registro = registro['fecha_registro']

        # Agregar los valores a la hoja
        hoja.append((titulo_pelicula, categoria_pelicula, ano_estreno, duracion_minutos, director, presupuesto, fecha_registro))

        # Iterar a través de las filas y aplicar el formato a la columna F
        for fila_num in range(2, hoja.max_row + 1):
            columna = 6  # Columna F
            celda = hoja.cell(row=fila_num, column=columna)
            celda.number_format = formato_moneda

    fecha_actual = datetime.datetime.now()
    archivoExcel = f"Reporte_peliculas_{fecha_actual.strftime('%Y_%m_%d')}.xlsx"
    carpeta_descarga = "../static/downloads-excel"
    ruta_descarga = os.path.join(os.path.dirname(
        os.path.abspath(__file__)), carpeta_descarga)

    if not os.path.exists(ruta_descarga):
        os.makedirs(ruta_descarga)
        # Dando permisos a la carpeta
        os.chmod(ruta_descarga, 0o755)

    ruta_archivo = os.path.join(ruta_descarga, archivoExcel)
    wb.save(ruta_archivo)

    # Enviar el archivo como respuesta HTTP
    return send_file(ruta_archivo, as_attachment=True)


def buscarPeliculaBD(search):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = """
                SELECT id_pelicula, titulo_pelicula, categoria_pelicula, ano_estreno, duracion_minutos, director, IFNULL(presupuesto, 0) AS presupuesto 
                FROM tbl_peliculas
                WHERE titulo_pelicula LIKE %s OR categoria_pelicula LIKE %s
                ORDER BY id_pelicula DESC
                """
                search_pattern = f"%{search}%"
                cursor.execute(querySQL, (search_pattern, search_pattern))
                peliculas = cursor.fetchall()
        return peliculas
    except Exception as e:
        print(f"Error al buscar películas: {e}")
        return []


def buscarPeliculaUnica(id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = ("""
                        SELECT 
                            p.id_pelicula,
                            p.titulo_pelicula,
                            p.categoria_pelicula,
                            p.ano_estreno,
                            p.duracion_minutos,
                            p.descripcion,
                            p.director,
                            IFNULL(p.presupuesto, 0) AS presupuesto,
                            p.poster_pelicula
                        FROM tbl_peliculas AS p
                        WHERE p.id_pelicula = %s LIMIT 1
                    """)
                mycursor.execute(querySQL, (id,))
                pelicula = mycursor.fetchone()
                return pelicula

    except Exception as e:
        print(f"Ocurrió un error en def buscarPeliculaUnica: {e}")
        return []



def procesar_actualizacion_form_pelicula(data):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                # Extraer y procesar datos del formulario
                titulo_pelicula = data.form['titulo_pelicula']
                categoria_pelicula = data.form['categoria_pelicula']
                ano_estreno = data.form['ano_estreno']
                duracion_minutos = data.form['duracion_minutos']
                descripcion = data.form['descripcion']
                director = data.form['director']

                # Procesar presupuesto eliminando caracteres no numéricos
                presupuesto_sin_puntos = re.sub('[^0-9]+', '', data.form['presupuesto'])
                presupuesto = int(presupuesto_sin_puntos)
                id_pelicula = data.form['id_pelicula']

                # Construir consulta SQL y parámetros dinámicamente
                query_base = """
                    UPDATE tbl_peliculas
                    SET 
                        titulo_pelicula = %s,
                        categoria_pelicula = %s,
                        ano_estreno = %s,
                        duracion_minutos = %s,
                        descripcion = %s,
                        director = %s,
                        presupuesto = %s
                """
                params = [
                    titulo_pelicula, categoria_pelicula, ano_estreno,
                    duracion_minutos, descripcion, director, presupuesto
                ]

                # Verificar si se subió un archivo de póster
                if 'poster_pelicula' in data.files and data.files['poster_pelicula'].filename != '':
                    file = data.files['poster_pelicula']
                    posterForm = procesar_imagen_poster(file)
                    query_base += ", poster_pelicula = %s"
                    params.append(posterForm)

                # Agregar condición WHERE
                query_base += " WHERE id_pelicula = %s"
                params.append(id_pelicula)

                # Ejecutar la consulta
                cursor.execute(query_base, params)
                conexion_MySQLdb.commit()

        return cursor.rowcount or []
    except Exception as e:
        print(f"Ocurrió un error en procesar_actualizacion_form: {e}")
        return None


# Lista de Usuarios creados
def lista_usuariosBD():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "SELECT id, name_surname, email_user, created_user FROM users"
                cursor.execute(querySQL,)
                usuariosBD = cursor.fetchall()
        return usuariosBD
    except Exception as e:
        print(f"Error en lista_usuariosBD : {e}")
        return []


# Eliminar uEmpleado
# Eliminar una película
def eliminarPelicula(idPelicula, poster_pelicula):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "DELETE FROM tbl_peliculas WHERE id_pelicula = %s"
                cursor.execute(querySQL, (idPelicula,))
                conexion_MySQLdb.commit()

                # Borrar el archivo del póster
                basepath = os.path.abspath(os.path.dirname(__file__))
                poster_path = os.path.join(basepath, '../static/posters_peliculas/', poster_pelicula)
                if os.path.exists(poster_path):
                    os.remove(poster_path)

                return cursor.rowcount
    except Exception as e:
        print(f"Error al eliminar la película: {e}")
        return None


# Eliminar usuario
def eliminarUsuario(id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "DELETE FROM users WHERE id=%s"
                cursor.execute(querySQL, (id,))
                conexion_MySQLdb.commit()
                resultado_eliminar = cursor.rowcount

        return resultado_eliminar
    except Exception as e:
        print(f"Error en eliminarUsuario : {e}")
        return []
